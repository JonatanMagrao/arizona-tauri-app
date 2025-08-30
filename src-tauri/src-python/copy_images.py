from pathlib import Path
from openpyxl import load_workbook
import time, os, json, shutil

def importar_produtos(lista_codigos, pasta_origem, pasta_destino):
    origem = Path(pasta_origem)
    destino = Path(pasta_destino)
    destino.mkdir(parents=True, exist_ok=True)

    arquivos = list(origem.iterdir())  
    imported_files = []
    not_found_files = []

    for codigo in lista_codigos:
        encontrados = [f for f in arquivos if codigo in f.stem]

        if encontrados:
            for arquivo in encontrados:
                try:
                    shutil.copy(arquivo, destino / arquivo.name)
                    imported_files.append(arquivo.name)
                except Exception as e:
                    raise Exception(f"⚠️ Erro ao copiar {arquivo.name}: {e}")
        else:
            not_found_files.append(codigo)

    return {
        "imported_files": imported_files,
        "not_found_files": not_found_files,
        "total_files": len(lista_codigos)
    }

def images_list_from_xl(xl_path, sheet_name):
  planilha = load_workbook(xl_path)
  aba = planilha[sheet_name]
  valores_unicos = set()

  for row in aba.iter_rows(values_only=True):
      if row[0] is not None:  # só processa se houver valor
          partes = [p.strip() for p in str(row[0]).split(";")]
          for texto in partes:
              valores_unicos.add(texto.split(".")[0])

  valores_unicos_lista = list(valores_unicos)
  return valores_unicos_lista


pasta_origem = r"J:\Drives compartilhados\Phx CRF\CARREFOUR\ASSETS\_FOTOS FLOW"
sheet_name = "Consolidado"

pasta_destino = r"copias"
sheet_path = r"C:\Users\PC\Downloads\Consolidado.xlsx"

lista_codigos = images_list_from_xl(sheet_path, sheet_name)
arquivos = importar_produtos(lista_codigos, pasta_origem, pasta_destino)
