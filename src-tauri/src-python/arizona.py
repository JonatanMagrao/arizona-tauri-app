# ================================================================
#  Criado por: Jonatan Magrão
#  Email:      magrao@jonatanmagrao.com.br
#  LinkedIn:   https://www.linkedin.com/in/jonatanmagrao/
#  Instagram:  @nerd_do_after
#
#  Direitos Autorais © 2025 - Todos os direitos reservados
#
#  ⚠️ Este código é fornecido exclusivamente para uso interno
#  da versão enviada por Jonatan Magrão. É proibida a alteração,
#  modificação, redistribuição ou utilização parcial/total sem
#  autorização prévia por escrito do autor.
# ================================================================

import os
import re
import json
import subprocess
import shutil
from openpyxl import load_workbook
from time import sleep
from datetime import datetime
from pathlib import Path


class Arizona:
    MONTH_NAMES_PT = {
        1: "JANEIRO", 2: "FEVEREIRO", 3: "MARCO", 4: "ABRIL",
        5: "MAIO", 6: "JUNHO", 7: "JULHO", 8: "AGOSTO",
        9: "SETEMBRO", 10: "OUTUBRO", 11: "NOVEMBRO", 12: "DEZEMBRO"
    }

    def __init__(self, config: dict, months_back=2):
        ae_version = config.get("ae_version")
        self.produtos = config.get("produtos", "PRODUTOS")
        drive_root = os.path.normpath(config.get("drive"))

        if not ae_version or not drive_root:
            raise ValueError(
                "Configuração inválida! É necessário ter 'ae_version' e 'drive'.")

        self.carrefour_path = os.path.join(drive_root, "Phx CRF")
        self.claro_path = os.path.join(
            drive_root, "Phx Talent", "CLARO", "2025")
        self.after_fx = os.path.normpath(
            rf"C:/Program Files/Adobe/Adobe After Effects {ae_version}/Support Files/AfterFX.exe")
        self.photoshop = os.path.normpath(
            r"C:/Program Files/Adobe/Adobe Photoshop 2024/Photoshop.exe")
        self.product_folder_path = os.path.join(
            rf"{self.carrefour_path}/CARREFOUR/ASSETS/_FOTOS FLOW")

        self.meses = self._build_month_labels(months_back)
        self.sheet_name = "Consolidado"

    def _build_month_labels(self, months_back: int):
        """
        Retorna lista como ['09_SETEMBRO', '08_AGOSTO', ...]
        começando do mês seguinte ao atual (se existir),
        caso contrário, começa no mês atual, e depois vai voltando.
        """
        today = datetime.today()
        cur_m = today.month
        labels = []

        # paths base para validar existência
        base_carrefour = os.path.join(
            self.carrefour_path, "CARREFOUR", "FILMES", "2025")

        # mês seguinte
        next_month = (cur_m % 12) + 1
        next_month_label = f"{next_month:02d}_{self.MONTH_NAMES_PT[next_month]}"
        next_month_path = os.path.join(base_carrefour, next_month_label)

        # verifica se existe pasta do mês seguinte
        if os.path.exists(next_month_path):
            labels.append(next_month_label)
        else:
            # fallback para mês atual
            cur_label = f"{cur_m:02d}_{self.MONTH_NAMES_PT[cur_m]}"
            labels.append(cur_label)

        # meses anteriores
        for i in range(months_back):
            m = ((cur_m - i - 1) % 12) + 1
            labels.append(f"{m:02d}_{self.MONTH_NAMES_PT[m]}")

        return labels

    def set_month_range(self, months_back: int):
        """Permite ajustar o range em runtime."""
        self.meses = self._build_month_labels(months_back)

    def open_visto(self):
        subprocess.run(
            ["start", "https://carrefour.visto.global/app/workspace/tasks"], shell=True)

    def open_bitrix(self):
        subprocess.run(
            ["start", "https://arizona.bitrix24.com/crm/type/1042/kanban/category/0/"], shell=True)

    def open_pip(self):
        subprocess.run(
            ["start", "https://cfo-pip.arizonaapps.io/site/jobs"], shell=True)

    def open_claro(self):
        subprocess.run(
            ["start", "https://talentmarcelclaro.visto.global/app/login"], shell=True)

    @property
    def open_produtos(self):
        subprocess.run(["explorer", os.path.join(
            self.carrefour_path, "CARREFOUR", "ASSETS", "_PRODUTOS")], shell=True)

    def open_out(self, jobao_cod, option):
        jobao_path = self.get_jobao_path(jobao_cod)
        options = {
            "mp4": os.path.join(jobao_path, "OUT", "RENDER", "MP4"),
            "mov": os.path.join(jobao_path, "OUT", "RENDER", "MOV"),
            "roteiro": os.path.join(jobao_path, "ROTEIRO", "LOCUCAO"),
            "print": os.path.join(jobao_path, "OUT", "PRINT"),
            "copia": os.path.join(jobao_path, "OUT", "COPIA"),
            "produtos": os.path.join(jobao_path, "PRODUTOS"),
            "claquetes": os.path.join(jobao_path, "CLAQUETES"),
            "audio": os.path.join(jobao_path, "AUDIO", "BOUNCE"),
        }

        try:
            subprocess.run(["explorer", options[option]], shell=True)
        except KeyError:
            raise Exception(f'Pasta "{option}" não encontrada em {jobao_cod}')

    def open_claro_folder(self):
        subprocess.run(
            ["explorer", "J:\\Drives compartilhados\\Phx Talent\\CLARO\\2025\\02_FEVEREIRO"], shell=True)

    def criar_novo_projeto_claro(self, claro_cod):
        MESES = self.meses
        for mes in MESES:
            projeto_path = os.path.join(self.claro_path, mes)
            if os.path.exists(projeto_path):
                novo_projeto_path = os.path.join(
                    projeto_path, f"CLARO - {claro_cod}")
                os.mkdir(novo_projeto_path)
                sleep(1)
                os.mkdir(os.path.join(novo_projeto_path, "OUT"))
                sleep(1)
                os.mkdir(os.path.join(novo_projeto_path, "PROJETO"))
                sleep(1)
                subprocess.run(["explorer", novo_projeto_path])
                return
        raise Exception(
            f"Não foi possível criar o projeto CLARO - {claro_cod} em {MESES[0]} nem em {MESES[1]}.")

    def open_tiago_folder(self):
        tiago_path = os.path.join(self.claro_path, "Tiago Leifert")
        subprocess.run(["explorer", tiago_path])

    def open_claro_projeto(self, claro_cod):
        MESES = self.meses
        for mes in MESES:
            projeto_path = os.path.join(
                self.claro_path, mes, f"CLARO - {claro_cod}", "PROJETO")
            if os.path.exists(projeto_path):
                subprocess.run(["explorer", projeto_path])
                return
        raise Exception(
            f'O projeto "{claro_cod}" não foi encontrado em {MESES[0]} nem em {MESES[1]}!')

    def open_claro_out(self, claro_cod):
        MESES = self.meses
        for mes in MESES:
            projeto_path = os.path.join(
                self.claro_path, mes, f"CLARO - {claro_cod}", "OUT")
            if os.path.exists(projeto_path):
                subprocess.run(["explorer", projeto_path])
                return
        raise Exception(
            f'O projeto "{claro_cod}" não foi encontrado em {MESES[0]} nem em {MESES[1]}!')

    def open_claro_aep(self, claro_cod):
        MESES = self.meses
        for mes in MESES:
            projeto_path = os.path.join(
                self.claro_path, mes, f"CLARO - {claro_cod}", "PROJETO")
            if os.path.exists(projeto_path):
                after_proj = [a for a in os.listdir(
                    projeto_path) if a.endswith(".aep")]
                if after_proj:
                    after_proj_path = os.path.join(projeto_path, after_proj[0])
                    subprocess.Popen(
                        [self.after_fx, "-project", after_proj_path])
                    return None
                else:
                    raise Exception(
                        f'Nenhum arquivo .aep foi encontrado no projeto "{claro_cod}".')
                    return None
        raise Exception(
            f'O projeto "{claro_cod}" não foi encontrado em {MESES[0]} nem em {MESES[1]}!')
        return None

    def open_claro_psd(self, claro_cod):
        MESES = self.meses
        for mes in MESES:
            projeto_path = os.path.join(
                self.claro_path, mes, f"CLARO - {claro_cod}")
            if os.path.exists(projeto_path):
                photoshop_proj = [a for a in os.listdir(
                    projeto_path) if a.endswith(".psd")]
                if photoshop_proj:
                    photoshop_path = os.path.join(
                        projeto_path, photoshop_proj[0])
                    subprocess.Popen([self.photoshop, photoshop_path])
                    return None
                else:
                    raise Exception(
                        f'Nenhum arquivo .psd foi encontrado no projeto "{claro_cod}".')
                    return None
        raise Exception(
            f'O projeto "{claro_cod}" não foi encontrado em {MESES[0]} nem em {MESES[1]}!')
        return None

    def get_jobao_path(self, jobao_cod):
        reg_exp = re.compile(rf'\d{{2}}_{jobao_cod}_\d{{5,6}}_w*')

        def get_directories(source):
            return [
                d for d in os.listdir(source)
                if os.path.isdir(os.path.join(source, d)) and re.match(r'\d{2}', d)
            ]

        MESES = self.meses
        for mes in MESES:
            projeto_path = os.path.join(
                self.carrefour_path, "CARREFOUR", "FILMES", "2025", mes)
            if not os.path.exists(projeto_path):
                continue
            pastas = get_directories(projeto_path)
            for pasta in pastas:
                if reg_exp.search(pasta):
                    return os.path.join(projeto_path, pasta)

        raise Exception(
            f'Jobão "{jobao_cod}" não encontrado em {MESES[0]} nem em {MESES[1]}!')

    def open_jobao(self, jobao_cod):
        jobao_path = self.get_jobao_path(jobao_cod)
        if jobao_path:
            subprocess.run(["explorer", jobao_path], shell=True)

    def find_spreadsheet(self, folder_path):
        folder = Path(folder_path)
        for file in folder.iterdir():
            if file.suffix == ".xlsx":
                return file

    def read_product_spreadsheet(self, jobao_cod):
        jobao_path = self.get_jobao_path(jobao_cod)
        produtos_folder = Path(jobao_path / "PRODUTOS")
        spreadsheet = self.find_spreadsheet(produtos_folder)

    def open_produtos_jobao(self, jobao_cod):
        jobao_path = self.get_jobao_path(jobao_cod)
        if jobao_path:
            subprocess.run(["explorer", os.path.join(
                jobao_path, "PRODUTOS")], shell=True)

    def open_jobinhos_folder(self, jobao_cod, jobinho_cod):
        jobao_path = os.path.join(
            self.get_jobao_path(jobao_cod), "PROJETOS", "AE")
        reg_exp = re.compile(rf'{jobinho_cod}_')
        arquivos = [a for a in os.listdir(jobao_path) if a.endswith(".aep")]
        for arquivo in arquivos:
            if reg_exp.search(arquivo):
                subprocess.run(["explorer", jobao_path], shell=True)
                return

    def abrir_jobinho(self, jobao_cod, jobinho_cod):
        jobao_path = os.path.join(
            self.get_jobao_path(jobao_cod), "PROJETOS", "AE")
        reg_exp = re.compile(rf'{jobinho_cod}_')
        arquivos = [a for a in os.listdir(jobao_path) if a.endswith(".aep")]
        for arquivo in arquivos:
            if reg_exp.search(arquivo):
                subprocess.Popen([self.after_fx, "-project",
                                 os.path.join(jobao_path, arquivo)])
                return
        raise Exception(f'Código Jobinho "{jobinho_cod}" inválido!')

    def find_prod(self, pesquisa):
        found_files = []
        for root, dirs, files in os.walk(os.path.join(self.carrefour_path, "CARREFOUR", "ASSETS", "_PRODUTOS")):
            for file in files:
                if ".psd" in file.lower() and pesquisa.lower() in file:
                    found_files.append(
                        {"nome": file, "path": "./icons/img.png"})
                    continue
                if pesquisa.lower() in file.lower():
                    found_files.append(
                        {"nome": file, "path": os.path.join(root, file)})
        return found_files if found_files else [{"nome": "Não existe no Banco de Dados", "path": "./icons/sad.png"}]

    def importar_produtos(self, dstn_folder, lista_codigos):
        origem = Path(self.product_folder_path)
        destino = Path(dstn_folder)

        arquivos = list(origem.iterdir())
        imported_files = []
        not_found_files = []

        for codigo in lista_codigos:
            codigo_lower = codigo.lower()

            # compara nomes ignorando maiúsculas/minúsculas
            encontrados = [
                f for f in arquivos
                if f.is_file() and f.stem.lower() == codigo_lower
            ]

            if encontrados:
                for arquivo in encontrados:
                    print(arquivo)
                    print(destino)
                    print(arquivo.name)
                    try:
                        shutil.copy(arquivo, destino / arquivo.name)
                        imported_files.append(arquivo.name)
                    except Exception as e:
                        raise Exception(
                            f"⚠️ Erro ao copiar {arquivo.name}: {e}")
            else:
                not_found_files.append(codigo)

        return {
            "imported_files": imported_files,
            "not_found_files": not_found_files,
            "total_files": len(lista_codigos)
        }

    def images_list_from_xl(self, jobao_cod):
        #! mudar o nome da pasta PRODUTOS_2 para PRODUTOS
        product_path = os.path.join(
            self.get_jobao_path(jobao_cod), self.produtos)
        for file in Path(product_path).iterdir():
            if file.is_file() and file.suffix.lower() == ".xlsx":
                sheet_path = str(file)   # garante string para load_workbook
                break

        if not sheet_path:
            raise FileNotFoundError(
                f"Nenhum .xlsx encontrado em {product_path}")
        planilha = load_workbook(sheet_path)
        aba = planilha[self.sheet_name]
        valores_unicos = set()

        for row in aba.iter_rows(values_only=True):
            if row[0] is not None:  # só processa se houver valor
                partes = [p.strip() for p in str(row[0]).split(";")]
                for texto in partes:
                    valores_unicos.add(texto.split(".")[0])

        valores_unicos_lista = list(valores_unicos)
        return product_path, valores_unicos_lista
